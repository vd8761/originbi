#!/usr/bin/env python3
"""
Generate the Level 3 "Metaphor" question import SQL from the "Level 3" sheet.

Usage:
    python generate_metaphor_sql.py "<Level 3.xlsx>" <output.sql>

Sheet columns: set number, question number, context_text_en, context_text_ta,
question_text_en, question_text_ta, image_desc_en, image_desc_ta, metadata.
(No image URL column — image_url is left NULL; Aarya provides URLs later.)

Idempotent via metadata->>'source' = 'metaphor_v1'. Open/global bank (no program scope).
Never touches a DB — only writes a .sql file.
"""
import sys
import json
import openpyxl

SHEET = "Level 3"
SOURCE_MARKER = "metaphor_v1"

# 0-based column indices
C_SET, C_QNO, C_CTX_EN, C_CTX_TA, C_Q_EN, C_Q_TA, C_IMG_EN, C_IMG_TA, C_META = range(9)


def sql_str(val):
    if val is None:
        return "NULL"
    s = str(val).strip()
    return "NULL" if s == "" else "'" + s.replace("'", "''") + "'"


def sql_jsonb(obj):
    return "'" + json.dumps(obj, ensure_ascii=False).replace("'", "''") + "'::jsonb"


def main():
    if len(sys.argv) < 3:
        print("usage: generate_metaphor_sql.py <xlsx> <out.sql>")
        sys.exit(1)
    xlsx, out = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET}' not found. Sheets: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    data = rows[1:]

    L = []
    out_lines = []
    add = out_lines.append
    set_counts = {}

    add("-- ============================================================")
    add("-- Level 3 Metaphor question import (auto-generated)")
    add(f"-- Source: {xlsx} [sheet: {SHEET}]")
    add(f"-- marker: {SOURCE_MARKER} | image_url left NULL (URLs added later)")
    add("-- Requires migration 015 (metaphor_questions).")
    add("-- Re-runnable: preamble removes any prior rows with this marker.")
    add("-- ============================================================")
    add("BEGIN;")
    add("")
    add(f"DELETE FROM metaphor_questions WHERE metadata->>'source' = '{SOURCE_MARKER}';")
    add("")
    add("INSERT INTO metaphor_questions")
    add("  (set_number, question_number, image_url,")
    add("   image_description_en, image_description_ta,")
    add("   context_text_en, context_text_ta, question_text_en, question_text_ta,")
    add("   metadata, is_active, is_deleted)")
    add("VALUES")

    value_rows = []
    for r in data:
        set_no = int(float(r[C_SET]))
        q_no = int(float(r[C_QNO])) if r[C_QNO] is not None else "NULL"
        set_counts[set_no] = set_counts.get(set_no, 0) + 1
        try:
            meta = json.loads(r[C_META]) if r[C_META] else {}
        except Exception:
            meta = {}
        meta["source"] = SOURCE_MARKER
        value_rows.append(
            f"  ({set_no}, {q_no}, NULL,\n"
            f"   {sql_str(r[C_IMG_EN])}, {sql_str(r[C_IMG_TA])},\n"
            f"   {sql_str(r[C_CTX_EN])}, {sql_str(r[C_CTX_TA])}, "
            f"{sql_str(r[C_Q_EN])}, {sql_str(r[C_Q_TA])},\n"
            f"   {sql_jsonb(meta)}, true, false)"
        )
    add(",\n".join(value_rows) + ";")
    add("")
    add(f"-- questions = {len(data)}, sets = {sorted(set_counts.items())}")
    add("COMMIT;")

    with open(out, "w", encoding="utf-8") as fh:
        fh.write("\n".join(out_lines))

    print(f"Wrote {out}")
    print(f"  metaphor questions: {len(data)}, per-set: {dict(sorted(set_counts.items()))}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Generate the SURVEY (open_questions) import SQL from the "Survey Questions"
sheet of the MBA workbook.

Usage:
    python generate_mba_survey_sql.py "<MBA Questions.xlsx>" <output.sql>

Mapping (Survey Questions sheet):
    set number      -> open_questions.set_number
    context_text_en -> open_questions.context_text_en (optional; blank -> NULL)
    question_text_en-> open_questions.question_text_en
    context_text_ta -> open_questions.context_text_ta (optional)
    question_text_ta-> open_questions.question_text_ta
    metadata        -> merged into open_questions.metadata (+ source marker)
    option A..D en/ta -> open_question_options (option_type 'TEXT', is_valid false)

All survey rows get question_type = 'SURVEY', media_type = 'TEXT'. Survey is
non-scoring, so options carry no DISC factor and is_valid = false.

Idempotent (preamble deletes prior rows with the source marker). Open questions
are global (not program/level scoped), so no sub-selects are needed.
This script never touches a database - it only writes a .sql file.
"""
import sys
import json
import openpyxl

SHEET = "Survey Questions"
QUESTION_TYPE = "SURVEY"
MEDIA_TYPE = "TEXT"
OPTION_TYPE = "TEXT"
SOURCE_MARKER = "mba_survey_v1"

# 0-based column indices in the Survey Questions sheet.
C_SET, C_CTX_EN, C_Q_EN, C_CTX_TA, C_Q_TA, C_META = 0, 1, 2, 3, 4, 5
OPTS = [(6, 7), (8, 9), (10, 11), (12, 13)]  # (en, ta) for A,B,C,D


def sql_str(val):
    if val is None:
        return "NULL"
    s = str(val).strip()
    if s == "":
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def sql_jsonb(obj):
    return "'" + json.dumps(obj, ensure_ascii=False).replace("'", "''") + "'::jsonb"


def main():
    if len(sys.argv) < 3:
        print("usage: generate_mba_survey_sql.py <xlsx> <out.sql>")
        sys.exit(1)
    xlsx, out = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET}' not found. Sheets: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    data = rows[1:]

    lines, L = [], None
    out_lines = []
    L = out_lines.append
    q_count = 0
    o_count = 0
    set_counts = {}

    L("-- ============================================================")
    L("-- MBA SURVEY open-question import (auto-generated)")
    L(f"-- Source: {xlsx} [sheet: {SHEET}]")
    L(f"-- question_type = {QUESTION_TYPE} | marker: {SOURCE_MARKER}")
    L("-- Re-runnable: preamble removes any prior rows with this marker.")
    L("-- Requires migration 014 (open_questions.set_number/context_text_*).")
    L("-- ============================================================")
    L("BEGIN;")
    L("")
    L("-- Idempotency: clear any previous SURVEY import of THIS batch.")
    L("DELETE FROM open_question_options o")
    L(" USING open_questions q")
    L(f" WHERE o.open_question_id = q.id AND q.metadata->>'source' = '{SOURCE_MARKER}';")
    L(f"DELETE FROM open_questions WHERE metadata->>'source' = '{SOURCE_MARKER}';")
    L("")

    for ri, r in enumerate(data, start=1):
        set_no = int(float(r[C_SET]))
        set_counts[set_no] = set_counts.get(set_no, 0) + 1
        try:
            meta = json.loads(r[C_META]) if r[C_META] else {}
        except Exception:
            meta = {}
        meta["source"] = SOURCE_MARKER

        q_count += 1
        L(f"-- Survey set {set_no} #{ri}")
        L("WITH oq AS (")
        L("  INSERT INTO open_questions")
        L("    (question_type, media_type, set_number,")
        L("     context_text_en, question_text_en, context_text_ta, question_text_ta,")
        L("     metadata, is_active, is_deleted)")
        L(f"  VALUES ('{QUESTION_TYPE}', '{MEDIA_TYPE}', {set_no},")
        L(f"     {sql_str(r[C_CTX_EN])}, {sql_str(r[C_Q_EN])}, "
          f"{sql_str(r[C_CTX_TA])}, {sql_str(r[C_Q_TA])},")
        L(f"     {sql_jsonb(meta)}, true, false)")
        L("  RETURNING id")
        L(")")
        L("INSERT INTO open_question_options")
        L("  (open_question_id, option_type, option_text_en, option_text_ta,")
        L("   is_valid, display_order, is_active, is_deleted)")
        L("SELECT oq.id, '" + OPTION_TYPE + "', v.en, v.ta, false, v.ord, true, false")
        L("FROM oq, (VALUES")
        opt_vals = []
        for oi, (cen, cta) in enumerate(OPTS, start=1):
            opt_vals.append(f"  ({oi}, {sql_str(r[cen])}, {sql_str(r[cta])})")
            o_count += 1
        L(",\n".join(opt_vals))
        L(") AS v(ord, en, ta);")
        L("")

    L(f"-- questions = {q_count}, options = {o_count}, sets = {sorted(set_counts.items())}")
    L("COMMIT;")

    with open(out, "w", encoding="utf-8") as fh:
        fh.write("\n".join(out_lines))

    print(f"Wrote {out}")
    print(f"  survey questions: {q_count}, options: {o_count}")
    print(f"  per-set counts: {dict(sorted(set_counts.items()))}")


if __name__ == "__main__":
    main()

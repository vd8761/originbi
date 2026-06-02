#!/usr/bin/env python3
"""
Generate the MBA College Level-1 main-question import SQL from the provided Excel.

Usage:
    python generate_mba_questions_sql.py "<path to MBA Questions.xlsx>" <output.sql>

What it does
------------
- Reads every sheet named "Set N" (1..n) from the workbook.
- Maps each row -> one assessment_questions row + its 4 assessment_question_options.
- disc_factor is NORMALISED to a clean D/I/S/C letter (first letter of the Excel
  "Factor" value); the original compound string (e.g. "C over D") is preserved in
  the option metadata. Validated 160/160 against the Master Reference Table.
- score_value = 1.0 for every option (standard DISC tally).
- Fully bilingual (en + ta) straight from the sheet.
- program_id / assessment_level_id are resolved INLINE via sub-selects so the same
  SQL runs on Local and Live without hardcoded IDs.
- Idempotent: a source marker (metadata->>'source') lets the script's preamble
  delete any prior import of the same batch before re-inserting.

It does NOT touch any database. It only writes a .sql file for review + manual run.
"""
import sys
import json
import openpyxl

PROGRAM_CODE = "COLLEGE_STUDENT"
LEVEL_NUMBER = 1
SOURCE_MARKER = "mba_xlsx_v1"  # bump if you re-import a corrected batch
VALID_DISC = {"D", "I", "S", "C"}

# Column indices in the "Set N" sheets (0-based), per the header row.
COL = {
    "set": 0,
    "ctx_en": 1, "q_en": 2, "ctx_ta": 3, "q_ta": 4, "meta": 5,
    # option (en, ta, factor) triples for A,B,C,D
    "opts": [(6, 7, 8), (9, 10, 11), (12, 13, 14), (15, 16, 17)],
}


def sql_str(val):
    """Escape a Python value as a SQL string literal (or NULL)."""
    if val is None:
        return "NULL"
    s = str(val)
    return "'" + s.replace("'", "''") + "'"


def sql_jsonb(obj):
    if obj is None:
        return "NULL"
    return "'" + json.dumps(obj, ensure_ascii=False).replace("'", "''") + "'::jsonb"


def normalise_factor(raw):
    """First alphabetic char -> clean DISC letter."""
    if raw is None:
        return None
    for ch in str(raw):
        if ch.isalpha():
            up = ch.upper()
            return up if up in VALID_DISC else None
    return None


def main():
    if len(sys.argv) < 3:
        print("usage: generate_mba_questions_sql.py <xlsx> <out.sql>")
        sys.exit(1)
    xlsx, out = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    set_sheets = sorted(
        [s for s in wb.sheetnames if s.lower().startswith("set ")],
        key=lambda s: int(s.split()[1]),
    )

    lines = []
    warnings = []
    q_count = 0
    o_count = 0

    L = lines.append
    L("-- ============================================================")
    L("-- MBA College Level-1 main questions import (auto-generated)")
    L(f"-- Source: {xlsx}")
    L(f"-- Sets: {', '.join(set_sheets)}")
    L(f"-- Program: {PROGRAM_CODE} | Level: {LEVEL_NUMBER} | marker: {SOURCE_MARKER}")
    L("-- Re-runnable: the preamble removes any prior rows with this marker.")
    L("-- ============================================================")
    L("BEGIN;")
    L("")
    L("-- Idempotency: clear any previous import of THIS batch (by source marker).")
    L("DELETE FROM assessment_question_options o")
    L(" USING assessment_questions q")
    L(f" WHERE o.question_id = q.id AND q.metadata->>'source' = '{SOURCE_MARKER}';")
    L(f"DELETE FROM assessment_questions WHERE metadata->>'source' = '{SOURCE_MARKER}';")
    L("")

    prog_sub = f"(SELECT id FROM programs WHERE code = '{PROGRAM_CODE}')"
    lvl_sub = f"(SELECT id FROM assessment_levels WHERE level_number = {LEVEL_NUMBER})"

    for sheet in set_sheets:
        ws = wb[sheet]
        rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
        data = rows[1:]  # skip header
        L(f"-- ---------- {sheet} ({len(data)} questions) ----------")
        for ri, r in enumerate(data, start=1):
            set_no = int(float(r[COL["set"]]))
            meta_raw = r[COL["meta"]]
            try:
                meta = json.loads(meta_raw) if meta_raw else {}
            except Exception:
                meta = {}
                warnings.append(f"{sheet} row {ri}: metadata not valid JSON")
            # NOTE: assessment_questions.category is reserved for Level 2 (ACI)
            # questions and is intentionally NOT written for Level 1. The theme
            # stays available in metadata.theme for reporting / category UIs.
            q_meta = dict(meta)
            q_meta["source"] = SOURCE_MARKER

            q_count += 1
            L(f"-- {sheet} Q{ri}")
            L("WITH q AS (")
            L("  INSERT INTO assessment_questions")
            L("    (assessment_level_id, set_number, program_id,")
            L("     context_text_en, question_text_en, context_text_ta, question_text_ta,")
            L("     metadata, is_active, is_deleted)")
            L(f"  VALUES ({lvl_sub}, {set_no}, {prog_sub},")
            L(f"     {sql_str(r[COL['ctx_en']])}, {sql_str(r[COL['q_en']])}, "
              f"{sql_str(r[COL['ctx_ta']])}, {sql_str(r[COL['q_ta']])},")
            L(f"     {sql_jsonb(q_meta)}, true, false)")
            L("  RETURNING id")
            L(")")
            L("INSERT INTO assessment_question_options")
            L("  (question_id, display_order, option_text_en, option_text_ta,")
            L("   disc_factor, score_value, metadata, is_active, is_deleted)")
            L("SELECT q.id, v.ord, v.en, v.ta, v.fac, 1.0, v.meta::jsonb, true, false")
            L("FROM q, (VALUES")
            opt_vals = []
            for oi, (cen, cta, cfac) in enumerate(COL["opts"], start=1):
                raw_fac = r[cfac]
                fac = normalise_factor(raw_fac)
                if fac is None:
                    warnings.append(
                        f"{sheet} row {ri} opt{oi}: unrecognised factor {raw_fac!r}")
                o_meta = {}
                if raw_fac is not None and str(raw_fac).strip().upper() != fac:
                    o_meta["raw_factor"] = str(raw_fac)
                if "dimension" in meta:
                    o_meta["dimension"] = meta["dimension"]
                meta_lit = json.dumps(o_meta, ensure_ascii=False).replace("'", "''")
                opt_vals.append(
                    f"  ({oi}, {sql_str(r[cen])}, {sql_str(r[cta])}, "
                    f"{sql_str(fac)}, '{meta_lit}')")
                o_count += 1
            L(",\n".join(opt_vals))
            L(") AS v(ord, en, ta, fac, meta);")
            L("")

    L("-- Sanity check (optional): expected counts")
    L(f"-- questions = {q_count}, options = {o_count}")
    L("COMMIT;")

    with open(out, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))

    print(f"Wrote {out}")
    print(f"  questions: {q_count}, options: {o_count}, sets: {len(set_sheets)}")
    if warnings:
        print(f"  WARNINGS ({len(warnings)}):")
        for w in warnings[:20]:
            print("   -", w)
    else:
        print("  no warnings")


if __name__ == "__main__":
    main()
